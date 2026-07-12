#!/usr/bin/env python3
"""
Stored Procedure Analyzer v5
- Physical tables ONLY — temp (#) and CTEs fully excluded from all output
- Alias collision detection: same alias name used for different tables in a single statement
- Alias scoped per-statement, not globally
- Single-table, no-alias SELECT/UPDATE/INSERT: unqualified columns attached to that table
- Predicates (CASE/WHERE/ON, simple col = ...) also counted as column usage
- Schema Breakdown tab: Schema -> Tables -> Columns
- Deterministic regex engine, zero LLM
- LookupValue exception: keep unqualified tokens under table-not-determined behavior
"""

import re
import sys
import argparse
from pathlib import Path
from collections import defaultdict
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

C_HEADER_BG = "1F3864"
C_HEADER_FG = "FFFFFF"
C_SUBHDR_BG = "2E75B6"
C_SUBHDR_FG = "FFFFFF"
C_ALT_ROW   = "DCE6F1"
C_WHITE     = "FFFFFF"
C_SCHEMA_BG = "E2EFDA"
C_SCHEMA_FG = "375623"

OP_COLORS = {
    "SELECT":   "375623",
    "INSERT":   "833C00",
    "UPDATE":   "7F6000",
    "DELETE":   "C00000",
    "MERGE":    "3A3268",
    "TRUNCATE": "420000",
}

def thin_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def style_header(cell, bg=C_HEADER_BG, fg=C_HEADER_FG, size=10):
    cell.font = Font(bold=True, color=fg, name="Arial", size=size)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border()

def style_data(cell, row_idx):
    cell.fill = PatternFill("solid", fgColor=C_ALT_ROW if row_idx % 2 == 0 else C_WHITE)
    cell.font = Font(name="Arial", size=10)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell.border = thin_border()

def style_schema_group(cell):
    cell.fill = PatternFill("solid", fgColor=C_SCHEMA_BG)
    cell.font = Font(bold=True, color=C_SCHEMA_FG, name="Arial", size=11)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = thin_border()

def op_color(ops):
    for op in ["DELETE", "TRUNCATE", "MERGE", "UPDATE", "INSERT", "SELECT"]:
        if op in ops:
            return OP_COLORS.get(op, "404040")
    return "404040"

def strip_name(n):
    return re.sub(r"[\[\]\"`]", "", str(n)).strip()

def is_temp(name):
    return str(name).lstrip("[").startswith("#")

def is_var(name):
    return str(name).startswith("@")

def detect_dialect(sql: str) -> str:
    s = sql.upper()
    if re.search(r"DECLARE\s+@", s):
        return "T-SQL (SQL Server)"
    if re.search(r"LANGUAGE\s+PLPGSQL", s):
        return "PostgreSQL"
    if re.search(r"CREATE\s+OR\s+REPLACE", s):
        return "PostgreSQL" if "LANGUAGE" in s else "Oracle PL/SQL"
    if re.search(r"`\w+`", sql):
        return "MySQL"
    if re.search(r"\$\$", sql):
        return "PostgreSQL"
    return "Auto-detected"

def split_procedures(sql: str):
    pat = re.compile(
        r"(CREATE\s+(?:OR\s+REPLACE\s+)?(?:PROCEDURE|PROC|FUNCTION)\s+([\w\.\[\]\"` ]+))",
        re.IGNORECASE,
    )
    matches = list(pat.finditer(sql))
    if not matches:
        return [("UnnamedProcedure", sql)]
    procs = []
    for i, m in enumerate(matches):
        start = m.start()
        end = matches[i + 1].start() if i + 1 < len(matches) else len(sql)
        name = strip_name(m.group(2)).strip()
        procs.append((name, sql[start:end]))
    return procs

def clean_sql(sql: str) -> str:
    sql = re.sub(r"/\*.*?\*/", " ", sql, flags=re.DOTALL)
    sql = re.sub(r"--[^\n]*", " ", sql)
    sql = re.sub(r"\s+", " ", sql)
    return sql.strip()

STMT_KEYWORDS = re.compile(
    r"(?=\b(?:SELECT|INSERT|UPDATE|DELETE|MERGE|TRUNCATE)\b)",
    re.IGNORECASE,
)

def split_statements(sql: str):
    parts = STMT_KEYWORDS.split(sql)
    return [p.strip() for p in parts if p.strip()]

SKIP_WORDS = {
    "SELECT","WHERE","SET","ON","AND","OR","NOT","IN","AS","WITH","BY","HAVING",
    "UNION","ALL","DISTINCT","TOP","NULL","BEGIN","END","IF","ELSE","THEN","CASE",
    "WHEN","RETURN","DECLARE","PRINT","GO","USE","OUTPUT","DEFAULT","VALUES",
    "EXISTS","COALESCE","ISNULL","CAST","CONVERT","GETDATE","SYSDATETIME","NEWID",
    "NOLOCK","READPAST","UPDLOCK","ROWLOCK","TABLOCK","INTO","FROM","EXEC","EXECUTE",
    "PROCEDURE","PROC","FUNCTION","TRIGGER","VIEW","TABLE","INDEX","DATABASE","SCHEMA",
    "SCOPE_IDENTITY","OBJECT_ID","ISNUMERIC","ISDATE","LEN","LTRIM","RTRIM","UPPER",
    "LOWER","SUBSTRING","CHARINDEX","REPLACE","STUFF","DATEDIFF","DATEADD","YEAR",
    "MONTH","DAY","COUNT","SUM","AVG","MIN","MAX","ROW_NUMBER","RANK","DENSE_RANK",
    "NTILE","LAG","LEAD","OVER","PARTITION","INSERTED","DELETED","IDENTITY","CONSTRAINT",
    "PRIMARY","FOREIGN","KEY","NVARCHAR","VARCHAR","INT","BIGINT","DATETIME","DATE",
    "BIT","DECIMAL","FLOAT","CHAR","TEXT","UNIQUEIDENTIFIER","MONEY","SMALLINT","TINYINT",
    "IMAGE","OBJECT_NAME","DB_NAME","USER_NAME","SUSER_NAME","HOST_NAME","SYSTEM_USER",
    "QUOTENAME","BETWEEN","LIKE","IS","MATCHED","TARGET","SOURCE","USING","WHEN","THEN",
    "MATCHED","BY",
}

def parse_table_ref(raw: str):
    raw = strip_name(raw).strip()
    parts = [p for p in raw.split(".") if p]
    if len(parts) >= 2:
        schema = parts[-2].upper()
        base = parts[-1].upper()
        full = f"{schema}.{base}"
    else:
        schema = ""
        base = parts[0].upper() if parts else raw.upper()
        full = base
    return schema, base, full

def build_alias_map(stmt: str):
    alias_map = {}
    alias_issues = []
    pat = re.compile(
        r"(?:FROM|JOIN|UPDATE|MERGE\s+(?:INTO\s+)?)\s+"
        r"((?:[\w\[\]]+\.)*[\w\[\]]+)"
        r"\s+(?:AS\s+)?([A-Za-z_]\w*)"
        r"(?=\s|\(|$|ON\b|SET\b|USING\b)",
        re.IGNORECASE,
    )
    for m in pat.finditer(stmt):
        traw = strip_name(m.group(1))
        alias = strip_name(m.group(2)).upper()
        if alias in SKIP_WORDS or is_temp(traw) or is_var(traw):
            continue
        schema, base, full = parse_table_ref(traw)
        if alias == base:
            continue
        existing = alias_map.get(alias)
        if existing and existing != full:
            alias_issues.append({
                "alias": alias,
                "table_prev": existing,
                "table_new": full,
                "statement": stmt.strip(),
            })
        else:
            alias_map[alias] = full
    return alias_map, alias_issues

TABLE_OP_PATTERNS = [
    (r"\bFROM\s+((?:[\w\[\]]+\.)*[\w\[\]]+)", "SELECT"),
    (r"\bINNER\s+JOIN\s+((?:[\w\[\]]+\.)*[\w\[\]]+)", "SELECT"),
    (r"\bLEFT\s+(?:OUTER\s+)?JOIN\s+((?:[\w\[\]]+\.)*[\w\[\]]+)", "SELECT"),
    (r"\bRIGHT\s+(?:OUTER\s+)?JOIN\s+((?:[\w\[\]]+\.)*[\w\[\]]+)", "SELECT"),
    (r"\bFULL\s+(?:OUTER\s+)?JOIN\s+((?:[\w\[\]]+\.)*[\w\[\]]+)", "SELECT"),
    (r"\bCROSS\s+JOIN\s+((?:[\w\[\]]+\.)*[\w\[\]]+)", "SELECT"),
    (r"\bJOIN\s+((?:[\w\[\]]+\.)*[\w\[\]]+)", "SELECT"),
    (r"\bINSERT\s+INTO\s+((?:[\w\[\]]+\.)*[\w\[\]]+)", "INSERT"),
    (r"\bUPDATE\s+((?:[\w\[\]]+\.)*[\w\[\]]+)\s+SET\b", "UPDATE"),
    (r"\bDELETE\s+FROM\s+((?:[\w\[\]]+\.)*[\w\[\]]+)", "DELETE"),
    (r"\bMERGE\s+(?:INTO\s+)?((?:[\w\[\]]+\.)*[\w\[\]]+)", "MERGE"),
    (r"\bTRUNCATE\s+TABLE\s+((?:[\w\[\]]+\.)*[\w\[\]]+)", "TRUNCATE"),
    (r"\bUSING\s+((?:[\w\[\]]+\.)*[\w\[\]]+)", "SELECT"),
]

def extract_all(sql: str):
    clean = clean_sql(sql)

    cte_names = set()
    for m in re.finditer(r"\bWITH\b\s+([\w\[\]]+)\s+AS\s*\(", clean, re.IGNORECASE):
        cte_names.add(strip_name(m.group(1)).upper())
    for m in re.finditer(r",\s*([\w\[\]]+)\s+AS\s*\(", clean, re.IGNORECASE):
        cte_names.add(strip_name(m.group(1)).upper())

    physical = {}

    def register(full_key, schema, base, op):
        if full_key not in physical:
            physical[full_key] = {
                "schema": schema,
                "base": base,
                "ops": set(),
                "aliases": set(),
                "columns": set(),
            }
        physical[full_key]["ops"].add(op)

    statements = split_statements(clean)
    alias_issues = []

    for stmt in statements:
        alias_map, stmt_issues = build_alias_map(stmt)
        if stmt_issues:
            alias_issues.extend(stmt_issues)

        stmt_tables = set()
        stmt_unresolved = set()

        for alias, full_key in alias_map.items():
            if full_key in physical:
                physical[full_key]["aliases"].add(alias)

        for pat, op in TABLE_OP_PATTERNS:
            for m in re.finditer(pat, stmt, re.IGNORECASE):
                raw = m.group(1).strip()
                schema, base, full = parse_table_ref(raw)
                if (
                    base in SKIP_WORDS
                    or base in cte_names
                    or is_temp(raw)
                    or is_var(raw)
                    or len(base) < 2
                ):
                    continue
                register(full, schema, base, op)
                stmt_tables.add(full)
                for a, fk in alias_map.items():
                    if fk == full and full in physical:
                        physical[full]["aliases"].add(a)

        # qualified columns
        for m in re.finditer(r"\b([\w\[\]]+)\.([\w\[\]]+)\b", stmt):
            prefix = strip_name(m.group(1)).upper()
            col = strip_name(m.group(2)).upper()
            if (
                col in SKIP_WORDS
                or is_var(col)
                or not re.match(r"^[A-Z_][A-Z0-9_]*$", col)
            ):
                continue
            # skip schema.table patterns (CLASSIFICATION.LINKCLASSDFAPROT10SCH)
            is_schema_table_ref = any(
                info.get("schema") == prefix and info.get("base") == col
                for info in physical.values()
                if info.get("base") != "__UNRESOLVED__"
            )
            if is_schema_table_ref:
                continue
            target_key = alias_map.get(prefix)
            if not target_key:
                for k in physical:
                    if k == prefix or k.endswith("." + prefix):
                        target_key = k
                        break
            if target_key and target_key in physical:
                physical[target_key]["columns"].add(col)

        # SELECT list (unqualified)
        for block_m in re.finditer(r"\bSELECT\b(.*?)\bFROM\b", stmt, re.IGNORECASE | re.DOTALL):
            block = re.sub(r"\(.*?\)", "", block_m.group(1), flags=re.DOTALL)
            for token in re.findall(r"\[[^]]+\]|[^,\s]+", block):
                raw_token = token
                # Skip pure string literals like 'AGREEMENT', "AGREEMENT"
                if raw_token.strip().startswith("'") or raw_token.strip().startswith('"'):
                    continue
                token = strip_name(token.split(".")[-1]).upper()
                if (
                    token
                    and " " not in token
                    and token not in SKIP_WORDS
                    and token != "*"
                    and re.match(r"^[A-Z_][A-Z0-9_]*$", token)
                ):
                    stmt_unresolved.add(token)

        # SET col =
        for m in re.finditer(r"\bSET\s+([\w\[\]]+)\s*=", stmt, re.IGNORECASE):
            col = strip_name(m.group(1)).upper()
            if col not in SKIP_WORDS:
                stmt_unresolved.add(col)

        # INSERT col list
        for m in re.finditer(r"\bINSERT\s+INTO\s+[\w\.\[\]]+\s*\((.*?)\)", stmt, re.IGNORECASE | re.DOTALL):
            for c in m.group(1).split(","):
                c = strip_name(c).upper().strip()
                if c and re.match(r"^[A-Z_][A-Z0-9_]*$", c) and c not in SKIP_WORDS:
                    stmt_unresolved.add(c)

        # Predicate columns (simple col = ...)
        for m in re.finditer(r"\b([A-Za-z_][A-Za-z0-9_]*)\s*=", stmt, re.IGNORECASE):
            col = strip_name(m.group(1)).upper()
            if col not in SKIP_WORDS and not is_var(col):
                stmt_unresolved.add(col)

        # Attach unresolved tokens for this statement
        if stmt_unresolved:
            if len(stmt_tables) == 1:
                only_table = next(iter(stmt_tables))
                _, base, _ = parse_table_ref(only_table)
                # Special case: LookupValue statements keep unqualified tokens under table-not-determined
                if base == "LOOKUPVALUE":
                    physical.setdefault(
                        "__UNRESOLVED__",
                        {
                            "schema": "",
                            "base": "__UNRESOLVED__",
                            "ops": set(),
                            "aliases": set(),
                            "columns": set(),
                        },
                    )
                    physical["__UNRESOLVED__"]["columns"].update(stmt_unresolved)
                else:
                    physical.setdefault(only_table, {
                        "schema": parse_table_ref(only_table)[0],
                        "base": parse_table_ref(only_table)[1],
                        "ops": set(),
                        "aliases": set(),
                        "columns": set(),
                    })
                    physical[only_table]["columns"].update(stmt_unresolved)
            else:
                physical.setdefault(
                    "__UNRESOLVED__",
                    {
                        "schema": "",
                        "base": "__UNRESOLVED__",
                        "ops": set(),
                        "aliases": set(),
                        "columns": set(),
                    },
                )
                physical["__UNRESOLVED__"]["columns"].update(stmt_unresolved)

    # clean up unresolved bucket
    mapped = set()
    for k, info in physical.items():
        if k != "__UNRESOLVED__":
            mapped.update(info["columns"])
    if "__UNRESOLVED__" in physical:
        physical["__UNRESOLVED__"]["columns"] -= mapped
        physical["__UNRESOLVED__"]["columns"] -= SKIP_WORDS
        if not physical["__UNRESOLVED__"]["columns"]:
            del physical["__UNRESOLVED__"]

    dynamic = bool(re.search(r"EXEC\s*\(|EXECUTE\s*\(|sp_executesql", clean, re.IGNORECASE))

    return physical, dynamic, alias_issues

def build_excel(all_results, output_path):
    wb = Workbook()

    # Summary
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value = "Stored Procedure Analyzer v5 — Physical Tables Report"
    c.font = Font(bold=True, name="Arial", size=14, color=C_HEADER_FG)
    c.fill = PatternFill("solid", fgColor=C_HEADER_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 34

    ws.merge_cells("A2:G2")
    c = ws["A2"]
    c.value = (
        f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}    |    "
        f"Procedures: {len(all_results)}    |    "
        f"Note: Temp tables and CTEs excluded — physical schema objects only"
    )
    c.font = Font(italic=True, name="Arial", size=9, color="595959")
    c.alignment = Alignment(horizontal="center")

    hdrs = ["#", "Procedure Name", "Source File", "Dialect",
            "Physical Tables", "Columns Resolved", "Dynamic SQL?"]
    for ci, h in enumerate(hdrs, 1):
        style_header(ws.cell(row=4, column=ci, value=h))
    ws.row_dimensions[4].height = 22

    for ri, (pname, source, dialect, physical, dynamic, alias_issues) in enumerate(all_results, 1):
        phys = {k: v for k, v in physical.items() if k != "__UNRESOLVED__"}
        total_cols = sum(len(v["columns"]) for v in phys.values())
        vals = [
            ri,
            pname,
            Path(source).name if source != "stdin" else "pasted input",
            dialect,
            len(phys),
            total_cols,
            "WARN" if dynamic else "No",
        ]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=ri + 4, column=ci, value=v)
            style_data(c, ri)
        if dynamic:
            ws.cell(row=ri + 4, column=7).font = Font(
                bold=True, color="C00000", name="Arial", size=10
            )

    for i, w in enumerate([5, 35, 25, 20, 16, 18, 14], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Physical Tables
    wt = wb.create_sheet("Physical Tables")
    wt.sheet_view.showGridLines = False

    hdrs = ["Procedure", "Schema", "Table Name", "Full Reference",
            "Operation(s)", "Aliases Used"]
    for ci, h in enumerate(hdrs, 1):
        style_header(wt.cell(row=1, column=ci, value=h))

    row = 2
    for pname, source, dialect, physical, dynamic, alias_issues in all_results:
        for key, info in sorted(
            ((k, v) for k, v in physical.items() if k != "__UNRESOLVED__"),
            key=lambda x: x[0],
        ):
            ops = ", ".join(sorted(info["ops"]))
            aliases = ", ".join(sorted(info["aliases"])) or "—"
            full = f"{info['schema']}.{info['base']}" if info["schema"] else info["base"]
            vals = [pname, info["schema"] or "(none)", info["base"], full, ops, aliases]
            for ci, v in enumerate(vals, 1):
                c = wt.cell(row=row, column=ci, value=v)
                style_data(c, row)
            wt.cell(row=row, column=5).font = Font(
                bold=True, color=op_color(info["ops"]), name="Arial", size=10
            )
            row += 1

    for i, w in enumerate([32, 14, 26, 30, 26, 22], 1):
        wt.column_dimensions[get_column_letter(i)].width = w

    # Columns Detail
    wc = wb.create_sheet("Columns Detail")
    wc.sheet_view.showGridLines = False

    hdrs = ["Procedure", "Schema", "Table Name", "Column Name",
            "Resolved Via", "Operation(s)"]
    for ci, h in enumerate(hdrs, 1):
        style_header(wc.cell(row=1, column=ci, value=h))

    row = 2
    for pname, source, dialect, physical, dynamic, alias_issues in all_results:
        for key, info in sorted(
            ((k, v) for k, v in physical.items() if k != "__UNRESOLVED__"),
            key=lambda x: x[0],
        ):
            schema = info["schema"] or "(none)"
            base = info["base"]
            ops = ", ".join(sorted(info["ops"])) or "—"
            cols = sorted(info["columns"])

            if not cols:
                vals = [
                    pname,
                    schema,
                    base,
                    "(no columns resolved)",
                    "alias.col / table.col",
                    ops,
                ]
                for ci, v in enumerate(vals, 1):
                    c = wc.cell(row=row, column=ci, value=v)
                    style_data(c, row)
                    if ci == 4:
                        c.font = Font(
                            italic=True, color="808080", name="Arial", size=10
                        )
                row += 1
            else:
                for col in cols:
                    vals = [
                        pname,
                        schema,
                        base,
                        col,
                        "alias.col / table.col",
                        ops,
                    ]
                    for ci, v in enumerate(vals, 1):
                        c = wc.cell(row=row, column=ci, value=v)
                        style_data(c, row)
                    row += 1

        # unresolved bucket
        unres_info = physical.get("__UNRESOLVED__")
        if unres_info:
            for col in sorted(unres_info["columns"]):
                vals = [
                    pname,
                    "—",
                    "(table not determined)",
                    col,
                    "unqualified token",
                    "—",
                ]
                for ci, v in enumerate(vals, 1):
                    c = wc.cell(row=row, column=ci, value=v)
                    style_data(c, row)
                    if ci in [2, 3, 5, 6]:
                        c.font = Font(
                            italic=True, color="808080", name="Arial", size=10
                        )
                row += 1

    for i, w in enumerate([32, 14, 26, 28, 22, 22], 1):
        wc.column_dimensions[get_column_letter(i)].width = w

    # Schema Breakdown
    ws4 = wb.create_sheet("Schema Breakdown")
    ws4.sheet_view.showGridLines = False

    ws4.merge_cells("A1:E1")
    c = ws4["A1"]
    c.value = "Schema Breakdown — Physical Tables & Columns Grouped by Schema"
    c.font = Font(bold=True, name="Arial", size=13, color=C_HEADER_FG)
    c.fill = PatternFill("solid", fgColor=C_HEADER_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws4.row_dimensions[1].height = 30

    hdrs = [
        "Schema", "Table Name", "Column Name",
        "Operations on Table", "Found In Procedure(s)",
    ]
    for ci, h in enumerate(hdrs, 1):
        style_header(ws4.cell(row=2, column=ci, value=h))

    schema_map = defaultdict(
        lambda: defaultdict(
            lambda: {"columns": set(), "ops": set(), "procs": set()}
        )
    )

    for pname, source, dialect, physical, dynamic, alias_issues in all_results:
        for key, info in physical.items():
            if key == "__UNRESOLVED__":
                continue
            schema = info["schema"] or "(no schema)"
            base = info["base"]
            schema_map[schema][base]["columns"].update(info["columns"])
            schema_map[schema][base]["ops"].update(info["ops"])
            schema_map[schema][base]["procs"].add(pname)

    row = 3
    for schema in sorted(schema_map.keys()):
        ws4.merge_cells(f"A{row}:E{row}")
        c = ws4.cell(row=row, column=1, value=f"  SCHEMA:  {schema}")
        style_schema_group(c)
        ws4.row_dimensions[row].height = 20
        row += 1

        for tname in sorted(schema_map[schema].keys()):
            tinfo = schema_map[schema][tname]
            cols = sorted(tinfo["columns"])
            ops = ", ".join(sorted(tinfo["ops"]))
            procs = ", ".join(sorted(tinfo["procs"]))

            if not cols:
                vals = [schema, tname, "(no columns resolved)", ops, procs]
                for ci, v in enumerate(vals, 1):
                    c = ws4.cell(row=row, column=ci, value=v)
                    style_data(c, row)
                    if ci == 3:
                        c.font = Font(
                            italic=True, color="808080", name="Arial", size=10
                        )
                ws4.cell(row=row, column=4).font = Font(
                    bold=True, color=op_color(tinfo["ops"]), name="Arial", size=10
                )
                row += 1
            else:
                for i, col in enumerate(cols):
                    vals = [
                        schema if i == 0 else "",
                        tname if i == 0 else "",
                        col,
                        ops if i == 0 else "",
                        procs if i == 0 else "",
                    ]
                    for ci, v in enumerate(vals, 1):
                        c = ws4.cell(row=row, column=ci, value=v)
                        style_data(c, row)
                    if i == 0:
                        ws4.cell(row=row, column=4).font = Font(
                            bold=True, color=op_color(tinfo["ops"]), name="Arial", size=10
                        )
                    row += 1

    for i, w in enumerate([16, 26, 28, 26, 45], 1):
        ws4.column_dimensions[get_column_letter(i)].width = w

    # Legend
    wl = wb.create_sheet("Legend")
    wl.sheet_view.showGridLines = False
    wl.merge_cells("A1:D1")
    c = wl["A1"]
    c.value = "Operation Color Legend & Analyst Notes"
    style_header(c, size=12)
    wl.row_dimensions[1].height = 26

    items = [
        ("SELECT",   OP_COLORS["SELECT"],   "Data read — FROM, JOIN, USING clauses"),
        ("INSERT",   OP_COLORS["INSERT"],   "Data write — INSERT INTO"),
        ("UPDATE",   OP_COLORS["UPDATE"],   "Data modification — UPDATE ... SET"),
        ("DELETE",   OP_COLORS["DELETE"],   "Data removal — DELETE FROM"),
        ("MERGE",    OP_COLORS["MERGE"],    "Upsert — MERGE INTO ... USING"),
        ("TRUNCATE", OP_COLORS["TRUNCATE"], "Full table wipe — TRUNCATE TABLE"),
    ]
    notes = [
        ("", ""),
        ("Scope", "Physical schema objects only — #temp tables and CTEs fully excluded"),
        ("Alias Collision", "Duplicate alias (same name, different tables) collected in alias_issues list"),
        ("Unresolved Cols", "Column found in SQL but table could not be determined from context"),
        ("Dynamic SQL WARN", "EXEC() / sp_executesql contents cannot be statically analyzed — review manually"),
        ("Schema Breakdown", "Groups all physical tables by schema with columns and procedures that reference them"),
    ]

    for ri, (op, color, desc) in enumerate(items, 2):
        wl.cell(row=ri, column=1, value=op).font = Font(
            bold=True, color=color, name="Arial", size=11
        )
        wl.cell(row=ri, column=1).alignment = Alignment(horizontal="center")
        wl.cell(row=ri, column=2, value="●").font = Font(color=color, size=14)
        wl.cell(row=ri, column=3, value=desc).font = Font(name="Arial", size=10)

    for ri, (label, note) in enumerate(notes, len(items) + 3):
        if label:
            wl.cell(row=ri, column=1, value=label).font = Font(
                bold=True, color=C_SUBHDR_BG, name="Arial", size=10
            )
        wl.cell(row=ri, column=3, value=note).font = Font(
            italic=True, name="Arial", size=10, color="404040"
        )

    for i, w in enumerate([18, 5, 70], 1):
        wl.column_dimensions[get_column_letter(i)].width = w

    wb.save(output_path)
    print(f"\nExcel report saved -> {output_path}\n")

def process_sql(sql: str, source: str, forced_dialect: str):
    dialect = detect_dialect(sql)
    if forced_dialect:
        dialect = forced_dialect
    procs = split_procedures(sql)
    results = []
    for pname, body in procs:
        physical, dynamic, alias_issues = extract_all(body)
        phys = {k: v for k, v in physical.items() if k != "__UNRESOLVED__"}
        src = Path(source).name if source != "stdin" else "pasted"
        print(
            f"  [{src}] -> {pname}: {len(phys)} physical table(s), "
            f"dialect={dialect}"
            + (" WARN dynamic SQL" if dynamic else "")
            + (f" | alias_issues={len(alias_issues)}" if alias_issues else "")
        )
        results.append((pname, source, dialect, physical, dynamic, alias_issues))
    return results

def main():
    parser = argparse.ArgumentParser(description="SP Analyzer v5 -> Excel")
    parser.add_argument("files", nargs="*", help=".sql files to process")
    parser.add_argument("--dialect", default="", help="Force dialect label")
    parser.add_argument("--output", default="sp_analysis_v5.xlsx")
    args = parser.parse_args()

    all_results = []
    if args.files:
        for fp in args.files:
            sql = Path(fp).read_text(encoding="utf-8", errors="ignore")
            # try:
            #     sql = Path(fp).read_text(encoding="utf-8", errors="ignore")
            # except unicodeDecodeError:
            #     sql =Path(fp).read_text(encoding="cp1252")
            all_results.extend(process_sql(sql, fp, args.dialect))
    else:
        print("Paste your stored procedure(s) below.")
        print("Press Ctrl+D (Mac/Linux) or Ctrl+Z+Enter (Windows) when done:\n")
        lines = []
        try:
            while True:
                lines.append(input())
        except EOFError:
            pass
        sql = "\n".join(lines)
        if not sql.strip():
            print("No SQL provided. Exiting.")
            sys.exit(1)
        all_results.extend(process_sql(sql, "stdin", args.dialect))

    if not all_results:
        print("Nothing to process.")
        sys.exit(1)

    build_excel(all_results, args.output)

if __name__ == "__main__":
    main()